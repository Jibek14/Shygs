import warnings
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
warnings.filterwarnings("ignore", "\nPyarrow", DeprecationWarning)


def create_general_df(data, df, is_import):
    rows = []

    for fruit_name, fruit_data in data.items():
        header = list(fruit_data.keys())
        rows.append(header)
        main_values = [value[0] if isinstance(value, list) else value for value in fruit_data.values()]
        rows.append(main_values)
        for i in range(len(fruit_data['наименование'][1])):
            if is_import == 0:
                row = [fruit_data['наименование'][1][i], fruit_data['кол-во'][1][i], fruit_data['ВЕС'][1][i]]
                rows.append(row)
            else:
                row = [fruit_data['наименование'][1][i], fruit_data['кол-во'][1][i]]
                rows.append(row)
    new_df = pd.DataFrame(rows)
    df = pd.concat([df, new_df], ignore_index=True)
    return df


def apply_color(sheet, pink_color="FC0FC0", default_color="FFFFFF"):
    pink_fill = PatternFill(start_color=pink_color, end_color=pink_color, fill_type="solid")
    pink_keywords = {"наименование", "кол-во", "ВЕС", "тара", "отпр", "закуп", "хозяин", "сумма", "маржа", "чистый вес"}
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            if cell.value in pink_keywords:
                cell.fill = pink_fill
                if cell.value == "наименование":
                    sheet.column_dimensions[cell.column_letter].width = 18  # Adjust width for "наименование" cells


def general_df_to_excel(df, output_file):
    try:
        df.to_excel(output_file, index=False,engine='openpyxl')
        wb = load_workbook(output_file)
        sheet = wb.active
        current_date = datetime.now().strftime("%d.%m.%Y")
        sheet.cell(row=1, column=1, value=f"Дата: {current_date}")
        apply_color(sheet)
        wb.save(output_file)
        print(f"DataFrame successfully saved to {output_file}")
    except Exception as e:
        print(f"An error occurred: {str(e)}")
