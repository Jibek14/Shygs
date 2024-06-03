import warnings
import datetime
from Fruit import Fruit
from openpyxl.styles import PatternFill
import pandas as pd
warnings.filterwarnings("ignore", "\nPyarrow", DeprecationWarning)


file_name = "список_для_клиентов.xlsx"
customers = {'АА': 'Айман', 'А': 'Айгуль Баха', 'Ж': 'Алтын', 'Б': 'Биба', 'И': 'Айнура', '$': 'Оразбек', 'О': 'Оля',
             'Э': 'Эльвира', 'Р': 'Роза', 'В': 'Вера', 'Ш': 'Анар', 'Г': 'Гуля Ардак', '8': 'Галина'}
data_per_page = {}
date = f"{datetime.date.today().day}.{datetime.date.today().month}.{datetime.date.today().year}"


def apply_color(sheet,color_a="FFCF40",color_b="FFD700"):
    yellow_fill = PatternFill(start_color=color_a, end_color=color_b, fill_type="solid")
    for row in sheet.iter_rows(min_row=2, max_row=2, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.fill = yellow_fill


def get_rows(obj: Fruit, nick):
    fruit_row = {
        'кол-во': obj.nick_localized_data[nick]['quantity'],
        'наименование': obj.title,
        'тара': obj.nick_localized_data[nick]['tare'],
        'брутто': obj.nick_localized_data[nick]['weight'],
        'нетто': obj.nick_localized_data[nick]['net'],
        'цена': obj.price,
        'сумма': obj.nick_localized_data[nick]['amount']
    }
    return fruit_row


def save_to_excel(df_per_page):
    res = {}
    try:
        with pd.ExcelWriter(file_name) as writer:
            for page, df in df_per_page.items():
                df.to_excel(writer, sheet_name=page, index=False, startrow=1)
                apply_color(writer.sheets[page])
                res[page] = {
                    'сумма': df['сумма'].sum(),
                    'Вес,работа': round(df['брутто'].sum() * 15),
                    'Итог': round(df['сумма'].sum() + round(df['брутто'].sum() * 14)),
                    'Жол': round(18 * df['брутто'].sum())
                }
                last_row = len(df) + 1

                writer.sheets[page].cell(row=last_row + 5, column=4, value=df['брутто'].sum())
                writer.sheets[page].cell(row=last_row + 5, column=6, value='15')
                writer.sheets[page].cell(row=last_row+ 7, column=6, value='18')
                writer.sheets[page].cell(row=1, column=2, value=f"{page} {date}")
                writer.sheets[page].cell(row=1, column=1, value=df['брутто'].sum())
                writer.sheets[page].column_dimensions['B'].width = 20
            for page, value in res.items():
                for index, (inner_key, inner_value) in enumerate(value.items()):
                    sheet = writer.sheets[page]
                    last_row = sheet.max_row-3

                    get_total_data(writer=writer, page=page, row= last_row+index, column=2, value=inner_key)
                    get_total_data(writer=writer, page=page, row=last_row + index, column=7, value=inner_value)
            print(f"All data saved to список.xlsx")
    except Exception as e:
        print(f"Ошибка: {e}")


def get_total_data(writer, page, row, column, value):
    writer.sheets[page].cell(row=row, column=column, value=value)


def create_data_frame(dict_data):
    df_per_page = {}
    for page, data in dict_data.items():
        df_per_page[page] = pd.DataFrame(data)
    return df_per_page


def get_pages(obj: Fruit):
    data = obj.nick_localized_data.items()
    for key, value in data:
        if key in customers.keys():
            if customers[key] not in data_per_page:
                data_per_page[customers[key]] = []
            data_per_page[customers[key]].append(get_rows(obj, key))
    if not data_per_page:
        print("Нет данных для создания файла.Проверьте имя клиента!")
    else:
        df_per_page = create_data_frame(data_per_page)
        save_to_excel(df_per_page)
